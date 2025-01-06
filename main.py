import logging
from api import wallapop_api, llm_ai_api
import numpy as np
import openpyxl
import httpx
from datetime import datetime
import io
import os
from PIL import Image

sheet_header = ['Nombre', 'Precio', 'Confianza', 'Explicación', 'Link', 'Imágenes']

title_permitted_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'

def main():
    logging.info('Starting...')

    # Set up temp images directory
    if not os.path.exists('./temp-images'):
        os.mkdir('./temp-images')

    # Set up out directory
    if not os.path.exists('./out'):
        os.mkdir('./out')

    products_listing = wallapop_api.get_listing()
    if not products_listing:
        exit(1)
    
    products_listing_text = ''
    num_id = 0
    for item in products_listing:
        products_listing_text = products_listing_text + f'\n\nID:{num_id} | **{item["title"]}** Descripción: \n{item["description"]}€' # {item["price"]}'
        num_id += 1

    keywords_of_listing = llm_ai_api.get_keywords(products_listing_text)

    workbook = openpyxl.Workbook()
    workbook_filename = f"out/{datetime.now().strftime('%H.%M %d-%m-%Y')}.xlsx"
    images_filename = f"{datetime.now().strftime('%H.%M_%d-%m-%Y')}.jpg"
    sheet = workbook.active
    for keywords_num, keywords_product in enumerate(keywords_of_listing, 1):
        if False and input(f"¿Quieres buscar el producto \"{keywords_product['keywords']}\" y/n ") != 'y': # Remove False if you want to pick which products to search
            continue
        keywords = keywords_product['keywords']
        res = wallapop_api.search_keywords(keywords)
        if len(res) < 5:
            logging.info(f'Not enough products with keywords: "{keywords}", pass')
            continue
        num_get_products = 1
        while len(res) % 40 == 0 and num_get_products < 2:
            num_get_products += 1
            res = res + wallapop_api.search_keywords(keywords, 1)
        logging.info(f'Found {len(res)} products using keywords: "{keywords}"')

        # Filtramos para este producto
        products_filter = llm_ai_api.filter_product_from_list(keywords, res)
        if not products_filter:
            exit(1)
        products_list = [res[x['id']] for x in products_filter if x['included']]
        logging.info(f'Found {len(products_list)} products after filtering using keywords: "{keywords}"')

        # Obtenemos los productos en primer cuarto de la distribución de precios
        q1 = np.percentile([x['price'] for x in products_list], 25)
        products_list_selected = [x for x in products_list if x['price'] <= q1 and not x['flags']['reserved']]
        logging.info(f'Starting valoration of {len(products_list_selected)} products.')

        # Escribimos el header y editamos el título de la página
        sheet.title = f"{''.join(char for char in keywords if char in title_permitted_characters)} {np.mean([x['price'] for x in products_list]):.2f}€"
        sheet.append(sheet_header)
        
        # Valoramos cada producto
        for num_product, product in enumerate(products_list_selected):
            product = products_list_selected[num_product]
            images = [httpx.get(x['xsmall']) for x in product['images']]
            confidence = llm_ai_api.get_product_confidence(product, images)
            line = [product['title'], product['price'], confidence['confidence'], confidence['reason'], f'https://es.wallapop.com/item/{product["web_slug"]}']
            sheet.column_dimensions[openpyxl.utils.get_column_letter(len(line))].width = 50 # Set link column width
            sheet.append(line)
            
            for num_img, img in enumerate(images, len(line) + 1):
                sheet.row_dimensions[num_product + 2].height = 150

                # Some images can be on webp format and openxlsx doesn't work with that
                # So we have to save those images in a temp folder in jpeg format, also, doing it like this solves the error "I/O Operation on closed file"
                image_pil = Image.open(io.BytesIO(img.read()))
                image_pil = image_pil.convert('RGB')
                image_filename = f'temp-images/image_{num_product}_{num_img}_{images_filename}'
                image_pil.save(image_filename, format = 'jpeg')
                img = openpyxl.drawing.image.Image(image_filename)         

                # TODO make the pictures look good in the excel       
                max_height = int(150*1.35)
                max_width = 100
                ratio = min(max_width/img.width, max_height/img.height)
                img.height = int(img.height * ratio)
                img.width = int(img.width * ratio)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(num_img)].width = 50
                sheet.add_image(img, f'{openpyxl.utils.get_column_letter(num_img)}{num_product + 2}')
            workbook.save(workbook_filename)
            workbook.close()
        if keywords_num < len(keywords):
            sheet = workbook.create_sheet()
    workbook.close()

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(filename)s - %(message)s')
    main()